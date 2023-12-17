import openpyxl

# File --> Workbook-->Sheets-->Rows-->Cells

file = "C:\Users\Admin\Desktop\College"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active  # get active sheet form the excel , sheet = workbook["sheet1"]

for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value="Welcome"

sheet.cell(1,1).value = 123
sheet.cell(1,2).value = 'Smith'
sheet.cell(1,3).value = 'Engineer'

sheet.cell(2,1).value = 567
sheet.cell(2,2).value = 'Doctor'

workbook.save(file)