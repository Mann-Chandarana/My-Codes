import openpyxl

# File --> Workbook-->Sheets-->Rows-->Cells

file = "C:\\Users\\Admin\Desktop\\College\\Semester-04\\DBMS\\Parth DBMS\\Excel_Tables_Temp"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Company_Details"]

rows = sheet.max_row   ### Count the number of rows in a excel sheet
cols = sheet.max_column  ### Count the number of columns in a excel sheet

### Reading all the rows and columsn from excel sheet

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end='   ')
    print()

